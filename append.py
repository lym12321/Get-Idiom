from openpyxl import Workbook
from openpyxl import load_workbook
import requests
import re
import json

name="auto.xlsx" # 文件名，可修改
url="https://hanyu.baidu.com/s?ptype=zici&wd=" # 使用百度汉语作为数据源
headers={
    'User-Agent':'UCWEB/2.0 (MIDP-2.0; U; Adr 9.0.0) UCBrowser U2/1.0.0 Gecko/63.0 Firefox/63.0 iPhone/7.1 SearchCraft/2.8.2 baiduboxapp/3.2.5.10 BingWeb/9.1 ALiSearchApp/2.4'
} # 模拟手机访问用的ua

wb = Workbook()
wb = load_workbook(name)

def cc(cy,i):
    # 录入前查重
    # 没用查找，直接遍历
    while i:
        if ws['B'+str(i)].value==cy:
            return True
        i-=1
    return False

def delete(i):
    # 看起来很傻但是很好用的删除方法
    print("已删除："+ws['B'+str(i)].value)
    ws['A'+str(i)]=None
    ws['B'+str(i)]=None
    ws['C'+str(i)]=None
    ws['D'+str(i)]=None

def append(id,chengyu,ciyi,beizhu):
    # 看起来很傻但是很好用的插入方法
    ws['A'+str(id+2)]=id
    ws['B'+str(id+2)]=chengyu
    ws['C'+str(id+2)]=ciyi
    ws['D'+str(id+2)]=beizhu

def setup():
    # 初始化表格
    ws.merge_cells("A1:D1")
    ws['A1']="成语库"
    ws.append(["id","成语","词义","备注"])

if wb.sheetnames[0]!="auto":
    ws = wb.create_sheet("auto",0)
    setup()
else:
    ws = wb.active

i=2
while ws['B'+str(i)].value!=None: i+=1 # 看起来很傻但是很好用的统计方法
print("已录入",i-3,"个成语")
while 1:
    strin = str(input(">>"))
    if strin=="end": break
    if strin=="del":
        # 删除上一个成语
        i-=1
        delete(i)
        wb.save(name)
        continue
    if strin=="count":
        print("已录入",i-3,"个成语")
        continue
    if cc(strin,i-1):
        # 录入前需要查重
        print("已存在")
    else:
        res = requests.get(url=url+strin,headers=headers)
        #r = re.findall(r"【解释】：(.+?)</li><li>",res.text,re.S)
        r = re.findall(r"window.basicInfo = (.+?);",res.text,re.S)
        #print(r[0])
        if r:
            # 很生硬但是可以用的截取方法（doge）
            basicInfo = json.loads(r[0])
            last = basicInfo['definition']
            last = last.split("##")[1]
            print(last)
            append(i-2,strin,last,None) #备注一栏暂时是None|后期可以改为其他数据
            wb.save(name) # 保存文件
            i+=1
        else:
            #print(res.text)
            #若想查看返回信息需要把上一行的注释删掉
            print("Unknow")
